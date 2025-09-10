import pandas as pd
import openpyxl

def read_xks_excel(file_path):
    # Load workbook and sheet
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Extract specific cell values
    info = {
        'B2': ws['B2'].value,
        'B6': ws['B6'].value,
        'D6': ws['D6'].value
    }


    # Read the header (row 8, 1-based index)
    header_row = 8
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        headers.append(cell.value)

    # Read data starting from row 9
    data = []
    for row in ws.iter_rows(min_row=9, max_row=ws.max_row, max_col=ws.max_column):
        row_data = []
        col = 0
        while col < ws.max_column:
            # Handle merged columns for A-B, C-D, E-F
            if col in [0, 2, 4]:
                value = row[col].value
                row_data.append(value)
                col += 2  # Skip merged pair
            else:
                row_data.append(row[col].value)
                col += 1
        data.append(row_data)

    # Adjust headers for merged columns
    merged_headers = []
    col = 0
    while col < len(headers):
        if col in [0, 2, 4]:
            merged_headers.append(headers[col])
            col += 2
        else:
            merged_headers.append(headers[col])
            col += 1

    # Create DataFrame
    df = pd.DataFrame(data, columns=merged_headers)

    # Convert all columns except the first to numeric
    for col in df.columns[1:]:
        df[col] = pd.to_numeric(df[col], errors='coerce')

    return df, info




# Example usage:
""" df, info = read_xks_excel(r"C:\Code\TIP\Balance_hidrico\input\estaciones_ideam\1_El Paraiso.xlsx")
print(info)
print(df.head()) """